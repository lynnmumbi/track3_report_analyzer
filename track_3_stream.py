import openpyxl
from openpyxl.styles import  Border, Side
import pandas as pd
from openpyxl.formatting.rule import  IconSetRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from copy import copy
from io import BytesIO

def process_excel(file_path):
    # Loads the workbook
    wb = openpyxl.load_workbook(file_path)

    # Access the first sheet
    sheet = wb['Scoring']


    # Insert space for headers and descriptions
    sheet.insert_rows(1, amount=22)

    # Format the header section
    def format_header():
        sheet.merge_cells('A2:O2')
        sheet.merge_cells('A3:O3')
        sheet.cell(row=2, column=1).value = "# MONTHLY FLEET REPORT"
        sheet.cell(row=3, column=1).value = "MONTH"

        for row in [2, 3]:
            cell = sheet.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True, underline='single')

        sheet.merge_cells('A5:O5')
        sheet.merge_cells('A6:O6')
        sheet.cell(row=5, column=1).value = "RAG REPORT"
        for row in [5, 6]:
            cell = sheet.cell(row=row, column=1)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=True)

        sheet.merge_cells('A8:I8')
        cell = sheet.cell(row=8, column=1)
        cell.value = "This report shows a classification of drivers in 3 different categories: Green, Amber, and Red."
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        sheet.merge_cells('A9:P9')
        cell = sheet.cell(row=9, column=1)
        cell.value = "We have also made a comparison between the two months, The Red dots are an indication that a vehicle increased on the violations from the previous month, the Light green dots shows an improvement on the different drivers and amber shows no change."
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    from openpyxl.styles import PatternFill, Font, Alignment

    def add_category_description(start_row, name, description, color, sheet):
        # Merge cells for category name
        sheet.merge_cells(f'A{start_row}:D{start_row}')
        cell_name = sheet.cell(row=start_row, column=1)
        cell_name.value = name
        cell_name.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell_name.font = Font(bold=True)
        cell_name.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Merge cells for category description
        sheet.merge_cells(f'A{start_row + 1}:I{start_row + 1}')
        cell_desc = sheet.cell(row=start_row + 1, column=1)
        cell_desc.value = description
        cell_desc.alignment = Alignment(wrap_text=True, vertical="top")  # Wrap text enabled


    # Add header and category descriptions for Green, Amber, and Red
    format_header()

    # Green category (10 to 13)
    green_description = (
        "The drivers in this group can serve as mentors or coaches for the rest of the team."

    )
    add_category_description(10, "Green Drivers (0 - 20 violations)", green_description, "339933", sheet)

    # Amber category (15 to 17)
    amber_description = (
        "These are the average drivers who fall in the middle range of performance. While they are neither particularly good "
        "nor bad, we recommend they receive guidance from the top-performing (green) drivers to help them improve."
    )
    add_category_description(15, "Amber Drivers (21 - 40 violations)", amber_description, "FFC000", sheet)

    # Red category (19 to 22)
    red_description = (
        "These drivers require immediate coaching and support. We recommend pairing them with top-performing (green) drivers "
        "for mentorship, along with offering incentives to encourage improvement."
    )
    add_category_description(19, "Red Drivers (above 40 violations)", red_description, "FF0000", sheet)

    # Count and calculate the percentage for each category
    def get_category_counts_and_percentages():
        green_count = 0
        amber_count = 0
        red_count = 0
        total_vehicles = 0

        # Identify the Advanced Score column
        header_row = 23
        advanced_score_col = None

        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=header_row, column=col).value == "Advanced Score":
                advanced_score_col = col
                break

        if advanced_score_col:
            for row in range(header_row + 1, sheet.max_row + 1):
                try:
                    score = sheet.cell(row=row, column=advanced_score_col).value
                    score = int(score)
                    total_vehicles += 1

                    if score <= 20:
                        green_count += 1
                    elif 21 <= score <= 40:
                        amber_count += 1
                    else:
                        red_count += 1
                except (ValueError, TypeError):
                    continue  # Skip invalid scores

        # Calculate percentages
        green_percentage = (green_count / total_vehicles) * 100 if total_vehicles > 0 else 0
        amber_percentage = (amber_count / total_vehicles) * 100 if total_vehicles > 0 else 0
        red_percentage = (red_count / total_vehicles) * 100 if total_vehicles > 0 else 0

        return {
            "Green": {"count": green_count, "percentage": green_percentage},
            "Amber": {"count": amber_count, "percentage": amber_percentage},
            "Red": {"count": red_count, "percentage": red_percentage},
        }

    # Get category counts and percentages
    category_data = get_category_counts_and_percentages()

    # Insert calculated count and percentages
    sheet.merge_cells('A13:H13')
    sheet.cell(row=12, column=1).value = f"This category includes {category_data['Green']['count']} vehicles, accounting for {category_data['Green']['percentage']:.2f}% of the total fleet."
    sheet.merge_cells('A17:H17')
    sheet.cell(row=17, column=1).value = f"This category includes {category_data['Amber']['count']} vehicles, accounting for {category_data['Amber']['percentage']:.2f}% of the total fleet"
    sheet.merge_cells('A21:H21')
    sheet.cell(row=21, column=1).value = f"This category includes {category_data['Red']['count']} vehicles, accounting for {category_data['Red']['percentage']:.2f}% of the total fleet"

    # Add "The table below outlines the vehicles..." description to row 22
    sheet.merge_cells('A22:O22')
    sheet.cell(row=22, column=1).value = "The table below outlines the vehicles in the three categories:"
    sheet.cell(row=22, column=1).alignment = Alignment(horizontal='left', vertical='center')
    sheet.cell(row=22, column=1).font = Font(bold=True)


    #### adding columns to the table ###########33

    # Find the index of the "Advanced Score" column
    header_row = 23
    headers = [cell.value for cell in sheet[header_row]]
    advanced_score_index = headers.index("Advanced Score") + 1  # Convert to 1-based index


    advanced_score_col_letter = get_column_letter(advanced_score_index)  # Get column letter

    # Extract Data for Sorting (Starting from Row 24)
    data_rows = [row for row in sheet.iter_rows(min_row=24, max_row=sheet.max_row, values_only=True)]

    # Sort Rows by "Advanced Score" BEFORE modifying the table
    data_rows.sort(key=lambda x: x[advanced_score_index - 1] if isinstance(x[advanced_score_index - 1], (int, float)) else float('inf'))

    # Write Sorted Data Back
    for row_idx, row_data in enumerate(data_rows, start=24):
        for col_idx, value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=value)


    # Insert new columns before and after the "Advanced Score" column
    sheet.insert_cols(advanced_score_index)  # Insert BEFORE "Advanced Score"
    sheet.insert_cols(advanced_score_index + 2)  # Insert AFTER "Advanced Score"

    # Assign headers to the new columns
    sheet.cell(row=header_row, column=advanced_score_index).value = "Previous Month Advanced Score"
    sheet.cell(row=header_row, column=advanced_score_index + 2).value = "Advanced Score Change"




    # Format the table
    # Apply borders to the entire table (from row 23 onwards)
    table_start_row = 23
    for row in sheet.iter_rows(min_row=table_start_row, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )

    # Define the fill color (light blue)
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Define the header row number and number of columns
    header_row = 23
    num_columns = sheet.max_column

    # Apply formatting to each header cell
    for col in range(1, num_columns + 1):
        cell = sheet.cell(row=header_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = header_fill  # Apply light blue background
        cell.alignment = Alignment(wrap_text = True, horizontal= "center", vertical="center")

    # Optionally, auto-adjust column width based on content
    for col in range(1, num_columns + 1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].width = 20  # Adjust width as needed

    # Apply color coding to the Advanced Score column
    header_row = 23
    advanced_score_col = None

    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=header_row, column=col).value == "Advanced Score":
            advanced_score_col = col
            break

    # Define color fills
    green_fill = PatternFill(start_color="339933", end_color="00FF00", fill_type="solid")  # Green
    amber_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Amber
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

    # Apply color coding to the Advanced Score column
    if advanced_score_col:
        for row in range(header_row + 1, sheet.max_row + 1):
            try:
                score = sheet.cell(row=row, column=advanced_score_col).value
                score = int(score)
                cell = sheet.cell(row=row, column=advanced_score_col)

                if score <= 20:
                    cell.fill = green_fill
                elif 21 <= score <= 40:
                    cell.fill = amber_fill
                else:
                    cell.fill = red_fill
            except (ValueError, TypeError):
                continue  # Skip invalid scores



    ### THE VLOOKUP PART ###
    prev_month_sheet = wb["Previous_month"]  # This contains last month's scores name it exactly like this!!

    # Extract "Grouping" & "Advanced Score" from "Previous_month"
    prev_month_data = {}

    # Find column indexes in "Previous_month"
    prev_headers = [cell.value for cell in prev_month_sheet[1]]
    reg_num_index = prev_headers.index("Grouping") + 1
    adv_score_index = prev_headers.index("Advanced Score") + 1

    #Store values in a dictionary (like a lookup table)
    for row in prev_month_sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        reg_num = row[reg_num_index - 1]  # Adjust for 0-based index
        adv_score = row[adv_score_index - 1]
        prev_month_data[reg_num] = adv_score  # Store in dict

    # Step 2: Populate "Previous Month Advanced Score" in main table
    # Find column indexes in the main sheet
    main_headers = [cell.value for cell in sheet[23]]  # Header row is 23
    reg_num_col = main_headers.index("Grouping") + 1
    prev_month_col = main_headers.index("Previous Month Advanced Score") + 1  # Newly added column

    # Loop through rows & fill in previous month's scores
    for row in range(24, sheet.max_row + 1):  # Data starts from row 24
        reg_num = sheet.cell(row=row, column=reg_num_col).value  # Get registration number
        if reg_num in prev_month_data:
            sheet.cell(row=row, column=prev_month_col).value = prev_month_data[reg_num]  # VLOOKUP match


    #############################################################################
    # getting the difference of the two advanced scores
    # Find column indexes
    adv_score_col = main_headers.index("Advanced Score") + 1
    adv_score_change_col = main_headers.index("Advanced Score Change") + 1  # Newly added column

    # Loop through rows & calculate the difference
    for row in range(24, sheet.max_row + 1):  # Data starts from row 24
        prev_score = sheet.cell(row=row, column=prev_month_col).value  # Previous month score
        curr_score = sheet.cell(row=row, column=adv_score_col).value  # Current advanced score

        if prev_score is None or curr_score is None:  # If missing values
            sheet.cell(row=row, column=adv_score_change_col).value = "-"  # Replace with "-"
        else:
            sheet.cell(row=row, column=adv_score_change_col).value = curr_score - prev_score  # Compute difference


    ################# adding icon sets to advanced score change ####################

    # Identify the column letter for "Advanced Score Change"
    advanced_score_change_col = None  # Initialize variable

    for col in sheet.iter_cols(min_row=23, max_row=23):  # Loop through header row
        for cell in col:
            if cell.value == "Advanced Score Change":  # Find the correct column
                advanced_score_change_col = cell.column_letter  # Get letter (e.g., 'E')
                advanced_score_change_index = cell.column  # Get column index
                break  # Stop searching once found

    # Ensure column detection works
    if not advanced_score_change_col:
        raise ValueError("Column 'Advanced Score Change' not found!")

    # Convert column values to float to prevent Excel from treating them as text
    for row in sheet.iter_rows(min_row=24, max_row=sheet.max_row,
                               min_col=advanced_score_change_index, max_col=advanced_score_change_index):
        for cell in row:
            if isinstance(cell.value, str):  # If stored as text
                try:
                    cell.value = float(cell.value)  # Convert to number
                except ValueError:
                    cell.value = None  # Ensure missing values don't break formatting



    # Define the icon set rule with proper thresholds
    icon_rule = IconSetRule(
        icon_style="3TrafficLights1",  # Uses red, yellow, and green icons
        type="num",
        values=[-100000, 0, 1],
        showValue=True,
        reverse=True
    )






    #######################################################################################################################

    ############################ utilization #######################################

    #######################################################################################################################


    # Load workbook
    utilization_sheet = wb["Utilization"]

    # Shift the table down so that it starts from row 13
    # Identify the first row with actual data (assumed directly after headers)
    data_start_row = 1  # Adjust if needed
    data_end_row = utilization_sheet.max_row  # Get last used row

    # Calculate how many rows to shift (new start is row 13)
    shift_by = 13 - data_start_row


    if shift_by > 0:
        for row in range(data_end_row, data_start_row - 1, -1):  # Move from bottom to top
            for col in range(1, utilization_sheet.max_column + 1):
                old_cell = utilization_sheet.cell(row=row, column=col)
                new_cell = utilization_sheet.cell(row=row + shift_by, column=col)

                # Copy values and formatting
                new_cell.value = old_cell.value
                new_cell.font = copy(old_cell.font)
                new_cell.fill = copy(old_cell.fill)
                new_cell.alignment = copy(old_cell.alignment)
                new_cell.border = copy(old_cell.border)

                # Clear old cell
                old_cell.value = None

    print(f"Table successfully shifted down. Data now starts at row 13.")



    # Clear existing formatting in the first 12 rows
    for row in range(1, 13):
        for cell in utilization_sheet[row]:
            cell.font = Font(bold=False)
            cell.alignment = Alignment(horizontal="left")

    # Remove borders from column A, rows 1 to 10
    for row in range(1, 11):
        utilization_sheet[f"A{row}"].border = Border()


    # Merge row 1 from A to AJ
    merge_range = f"A1:AJ1"
    utilization_sheet.merge_cells(merge_range)

    # Remove borders from the first row
    for col in range(1, utilization_sheet.max_column + 1):
        utilization_sheet.cell(row=1, column=col).border = Border()


    # Format merged cell
    cell = utilization_sheet["A1"]
    cell.value = "DAILY UTILIZATION REPORT"
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.font = Font(bold=True, underline="single")

    #increasing row height
    utilization_sheet.row_dimensions[1].height = 25

    #leave out row 2 and 3
    for row in range(2,4):
        for cell in utilization_sheet[row]:
            cell.value= None

    # Define color fills
    red_fill_util = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    amber_fill_util = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
    yellow_fill_util = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    green_fill_util= PatternFill(start_color="008000", end_color="008000", fill_type="solid")  # Dark Green

    # Define border style
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Define descriptions
    descriptions = [
        "Less Than 0.1 km",
        "Less Than 10 km",
        "Less Than 100 Km",
        "More Than 100 Km"
    ]



    #color list
    color_fills = [red_fill_util, amber_fill_util, yellow_fill_util,green_fill_util]

    # Find "Weekday Distance (km)" column dynamically
    table_start_row_uti_col = 14  # Assuming headers start at row 13
    header_row_uti_number = 13
    weekday_distance_col_index = None

    for cell in utilization_sheet[header_row_uti_number]:
        if cell.value == "Weekday Distance (km)":
            weekday_distance_col_index = cell.column
            break

    if weekday_distance_col_index is None:
        raise ValueError("Could not find 'Weekday Distance (km)' column.")

    # Set the last column to color (just before "Weekday Distance (km)")
    table_start_col = 1
    table_end_col = weekday_distance_col_index - 1

    # Rows to color - start from row 14 (after headers), or your desired start row
    table_start_row = 14
    table_end_row = utilization_sheet.max_row 


    # Identify the column index for 'Total Distance (km)' (replace with the correct column name if different)
total_distance_col_index = None
for col in range(1, utilization_sheet.max_column + 1):
    if utilization_sheet.cell(row=13, column=col).value == "Total Distance (km)":
        total_distance_col_index = col
        break

# Ensure we found the column
if total_distance_col_index is None:
    raise ValueError("Column 'Total Distance (km)' not found.")

# Insert a new row for Totals
totals_row = table_end_row  + 1
utilization_sheet.insert_rows(totals_row)

# Label the first cell in the totals row
utilization_sheet.cell(row=totals_row, column=1, value="Totals").font = Font(bold=True)

# Sum each column from row 14 to the last data row
for col in range(2, total_distance_col_index + 1):  # Start from col 2 to exclude labels
    col_letter = get_column_letter(col)
    sum_formula = f"=SUM({col_letter}{table_start_row}:{col_letter}{table_end_row })"
    utilization_sheet.cell(row=totals_row, column=col, value=sum_formula).font = Font(bold=True)


    # Apply color coding based 

    for row in utilization_sheet.iter_rows(min_row=table_start_row, max_row=table_end_row,
                                           min_col=table_start_col, max_col=table_end_col):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                if cell.value >= 100:
                    cell.fill = green_fill_util
                elif cell.value >= 10:
                    cell.fill = yellow_fill_util
                elif cell.value >= 0.1:
                    cell.fill = amber_fill_util
                else:
                    cell.fill = red_fill_util


    # Apply colors, descriptions, and borders
    for i, desc in enumerate(descriptions):
        row = i + 4  # Row 4 to 7
        cell_q = utilization_sheet[f"Q{row}"]
        cell_q.fill = color_fills[i]  # Apply color
        cell_q.border = thin_border  # Apply border

        utilization_sheet.merge_cells(f"R{row}:V{row}")  # Merge R to V
        cell_rv = utilization_sheet[f"R{row}"]
        cell_rv.value = desc  # Add description
        cell_rv.alignment = Alignment(horizontal="left")  # Align left
        cell_rv.border = thin_border  # Apply border to merged cells

        # Debugging statements to check values
        print(f"Setting description at row {row}: {desc}")
        print(f"Cell before merge: {utilization_sheet[f'R{row}'].value}")
        print(f"Cell after merge: {utilization_sheet[f'R{row}'].value}")

        # Apply borders to merged range manually (since only the top-left cell gets a border by default)
        for col in range(ord("R"), ord("V") + 1):  # Iterate from column R to V
            utilization_sheet[f"{chr(col)}{row}"].border = thin_border



    # Merge cells from Q9 to AA11
    utilization_sheet.merge_cells("Q9:AA9")
    utilization_sheet.merge_cells("Q10:AA10")
    utilization_sheet.merge_cells("Q11:AA11")



    # Identify headers dynamically from row 13
    headers_utilization = [cell.value for cell in utilization_sheet[13] if cell.value]  # Remove None values

    # Get column count
    total_columns = len(headers_utilization)

    # Identify key columns
    vehicle_col = 1  # First column is vehicle column
    total_distance_col = total_columns - 2  # Third last column (always)

    # Find the first non-empty row (Headers must be on or after row 13)
    header_row_util = 13  # Starts checking from row 13
    while header_row_util <= utilization_sheet.max_row:
        headers_utilization = [cell.value for cell in utilization_sheet[header_row_util] if cell.value]  # Remove empty values
        if headers_utilization:  # Found non-empty row
            break
        header_row_util += 1
    else:
        raise ValueError("No headers found in the sheet.")


    # Identify first and last data rows
    table_start_row = 14  # Data starts after header
    table_end_row = utilization_sheet.max_row  # Last row with data

    # Convert to Excel column letter
    total_distance_col_letter = utilization_sheet.cell(row=13, column=total_distance_col).column_letter

    # Get all distances except the totals row
    fleet_distances = [
        (utilization_sheet[f"A{row}"].value, utilization_sheet[f"{total_distance_col_letter}{row}"].value or 0)
        for row in range(table_start_row, table_end_row)  # Excluding totals
    ]

    # Sort by distance to find least and most utilized vehicles
    fleet_distances_sorted = sorted(fleet_distances, key=lambda x: x[1])  # Sort by distance


    least_utilized_vehicle, least_distance = fleet_distances_sorted[0]
    most_utilized_vehicle, most_distance = fleet_distances_sorted[-1]

    # Compute fleet average, excluding totals row
    fleet_average = sum([dist for _, dist in fleet_distances]) / len(fleet_distances) if fleet_distances else 0

    # Merge cells from Q9 to AA11
    utilization_sheet.merge_cells("Q9:AA9")
    utilization_sheet.merge_cells("Q10:AA10")
    utilization_sheet.merge_cells("Q11:AA11")

    # Set values
    utilization_sheet["Q9"].value = f"The least utilized vehicle was {least_utilized_vehicle} with {least_distance} KM"
    utilization_sheet["Q10"].value = f"The most utilized vehicle was {most_utilized_vehicle} with {most_distance} KM"
    utilization_sheet["Q11"].value = f"The average distance covered by each vehicle in the fleet was {fleet_average:.1f} KM"



    # Format text: bold for key terms
    for row in range(9, 12):
        cell = utilization_sheet[f"Q{row}"]
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.font = Font(bold=True)

    # color coding the uti table
    #damn this alot of code
    #already defined the border style
    table_start_col = 1
    table_end_col_with_total_dis=utilization_sheet.max_column
    table_end_col = utilization_sheet.max_column - 5  # Exclude last 5 columns

    # Apply borders to all data cells
    for row in range(table_start_row, table_end_row + 1):
        for col in range(table_start_col, table_end_col_with_total_dis + 1):
            utilization_sheet.cell(row=row, column=col).border = thin_border

    # Wrap text in headers & color headers light blue
    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")


    for col in range(table_start_col, table_end_col_with_total_dis + 1):
        header_cell = utilization_sheet.cell(row=header_row_util, column=col)
        header_cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        header_cell.fill = header_fill
        header_cell.border = thin_border  # Ensure headers also have borders

    # Define color mapping
    color_mapping = {
        "Red": "FF0000",      # Less than 0.1 km
        "Amber": "FFA500",    # Less than 10 km
        "Yellow": "FFFF00",   # Less than 100 km
        "Green": "008000"     # More than 100 km
    }

    # Get the range for data (excluding totals and last 5 columns)
    for row in range(table_start_row, table_end_row ):  # Loop through data rows
        for col in range(table_start_col, table_end_col - 4):  # Exclude last 5 columns
            cell = utilization_sheet.cell(row=row, column=col)

            # Convert cell value to a number (if possible)
            try:
                cell_value = float(cell.value)  # Convert to float if possible
            except (TypeError, ValueError):
                continue  # Skip non-numeric or empty cells

            # Apply color coding based on conditions
            if cell_value < 0.1:
                cell.fill = PatternFill(start_color=color_mapping["Red"], fill_type="solid")
            elif cell_value < 10:
                cell.fill = PatternFill(start_color=color_mapping["Amber"], fill_type="solid")
            elif cell_value < 100:
                cell.fill = PatternFill(start_color=color_mapping["Yellow"], fill_type="solid")
            else:
                cell.fill = PatternFill(start_color=color_mapping["Green"], fill_type="solid")


    #############################sheets to remain###############################################
    # List of sheets to retain in order
    sheets_to_keep_and_order = ["Scoring", "Utilization"]


    # Iterate over a list to avoid modifying the workbook object while iterating
    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep_and_order:
            wb.remove(wb[sheet_name])

    #Reorder the remaining sheets
    for index, sheet_name in enumerate(sheets_to_keep_and_order):
        if sheet_name in wb.sheetnames:
            wb.move_sheet(sheet_name, offset=index - wb.sheetnames.index(sheet_name))





    print("✅ Go see what you did baby girl!")
    # Save to a BytesIO object instead of a file
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream


